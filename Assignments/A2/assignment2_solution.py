"""
CMSC 5718 Assignment 2: Asset Allocation
Student Name: [Your Name]
Student Number: 1155239333

This script computes all answers for Assignment 2.
"""

import numpy as np
import pandas as pd
from scipy.optimize import minimize
import matplotlib.pyplot as plt
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# PART 0: Load Data from Excel
# ============================================================================

print("=" * 80)
print("CMSC 5718 Assignment 2: Asset Allocation")
print("Student Number: 1155239333")
print("=" * 80)

wb = openpyxl.load_workbook('CMSC 5718 Assignment 2 parameters.xlsx', data_only=True)

# Load stock prices
ws_prices = wb['stock prices']
stock_codes = [267, 388, 700, 941, 981, 1801, 2319, 2899, 9888, 9988]
stock_names = ['Citic Limited', 'Hong Kong Exchanges', 'Tencent', 'China Mobile',
               'SMIC', 'Innovent Biologics', 'China Mengniu Dairy', 'Zijin Mining',
               'Baidu', 'Alibaba']

# Parse price data (rows 4 onwards have data, row 2 has stock codes, row 3 has "Dates")
dates = []
price_data = {code: [] for code in stock_codes}
hsi_prices = []

for i, row in enumerate(ws_prices.iter_rows(values_only=True), 1):
    if i < 4:  # Skip header rows
        continue
    row_list = list(row)
    date_val = row_list[1]
    if date_val is None:
        continue
    if hasattr(date_val, 'strftime'):
        dates.append(date_val)
    else:
        from datetime import datetime
        try:
            dates.append(datetime.strptime(str(date_val), '%Y-%m-%d'))
        except:
            continue
    
    for j, code in enumerate(stock_codes):
        price_data[code].append(float(row_list[j + 2]))
    hsi_prices.append(float(row_list[12]))

# Create DataFrame
price_df = pd.DataFrame(price_data, index=dates)
price_df['HSI'] = hsi_prices

print(f"\nData loaded: {len(dates)} trading days from {dates[0].strftime('%Y-%m-%d')} to {dates[-1].strftime('%Y-%m-%d')}")
print(f"Number of stocks: {len(stock_codes)}")
print(f"Stock codes: {stock_codes}")

# Load expected returns
ws_er = wb['Consensus Expected return']
expected_returns = {}
for row in ws_er.iter_rows(values_only=True):
    row_list = list(row)
    if row_list[2] is not None and isinstance(row_list[2], (int, float)):
        code = int(row_list[2])
        if code in stock_codes:
            expected_returns[code] = float(row_list[4])

rf = 0.029  # Risk-free rate

print(f"\nExpected returns:")
for code in stock_codes:
    print(f"  {code} ({stock_names[stock_codes.index(code)]}): {expected_returns[code]*100:.2f}%")
print(f"  Risk-free rate: {rf*100:.2f}%")

# ============================================================================
# QUESTION 1: Collecting Historical Data (12 marks)
# ============================================================================
print("\n" + "=" * 80)
print("QUESTION 1: Collecting Historical Data")
print("=" * 80)

# Compute daily returns: r_t = (P_t - P_{t-1}) / P_{t-1}
returns_df = price_df.pct_change().dropna()
n = len(returns_df)  # number of returns (n+1 prices)
print(f"\nNumber of returns (n): {n}")
print(f"Number of prices (n+1): {n + 1}")

# Q1(i): Annualized standard deviation of daily return
# Annualized s.d. = daily s.d. * sqrt(252)
# Using sample standard deviation (ddof=1)
trading_days = 252

daily_std = returns_df.std(ddof=1)
annualized_std = daily_std * np.sqrt(trading_days)

print("\n--- Q1(i): Annualized Standard Deviation ---")
print(f"{'Stock':<10} {'Ann. Std Dev':>15}")
print("-" * 30)
for code in stock_codes:
    print(f"  {code:<8} {annualized_std[code]:>14.6f}  ({annualized_std[code]*100:.4f}%)")
print(f"  {'HSI':<8} {annualized_std['HSI']:>14.6f}  ({annualized_std['HSI']*100:.4f}%)")

# Q1(ii): Annualized covariance between each pair of stocks
# Annualized covariance = daily covariance * 252
# Using sample covariance (ddof=1, which is default for DataFrame.cov())
stock_returns = returns_df[stock_codes]
daily_cov = stock_returns.cov()  # ddof=1 by default
annualized_cov = daily_cov * trading_days

print("\n--- Q1(ii): Annualized Covariance Matrix ---")
print("\nAnnualized Covariance Matrix (10x10):")
print(annualized_cov.to_string(float_format=lambda x: f"{x:.8f}"))

# ============================================================================
# QUESTION 2: Finding the Optimal Portfolio (82 marks)
# ============================================================================
print("\n" + "=" * 80)
print("QUESTION 2: Finding the Optimal Portfolio")
print("=" * 80)

# Setup
mu = np.array([expected_returns[code] for code in stock_codes])  # expected returns vector
Sigma = annualized_cov.values  # covariance matrix
n_stocks = len(stock_codes)

# ---- Q2(i)(a): Equal Weight (EW) Portfolio ----
print("\n--- Q2(i)(a): Equal Weight (EW) Portfolio ---")
w_ew = np.array([0.1] * n_stocks)

# Portfolio expected return: E[r_p] = w' * mu
ew_return = np.dot(w_ew, mu)
# Portfolio variance: sigma_p^2 = w' * Sigma * w
ew_var = np.dot(w_ew, np.dot(Sigma, w_ew))
ew_std = np.sqrt(ew_var)

print(f"  Weights: {w_ew}")
print(f"  Expected Return: {ew_return:.6f} ({ew_return*100:.4f}%)")
print(f"  Standard Deviation: {ew_std:.6f} ({ew_std*100:.4f}%)")

# ---- Q2(i)(b): Global Minimum Variance (MV) Portfolio (no short selling) ----
print("\n--- Q2(i)(b): Global Minimum Variance (MV) Portfolio ---")

def portfolio_variance(w, Sigma):
    return np.dot(w, np.dot(Sigma, w))

# Constraints: sum of weights = 1
constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
# Bounds: no short selling (0 <= w_i <= 1)
bounds = tuple((0, 1) for _ in range(n_stocks))
# Initial guess: equal weights
w0 = np.array([1.0 / n_stocks] * n_stocks)

result_mv = minimize(portfolio_variance, w0, args=(Sigma,),
                     method='SLSQP', bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-15, 'maxiter': 1000})

w_mv = result_mv.x
mv_var = portfolio_variance(w_mv, Sigma)
mv_std = np.sqrt(mv_var)
mv_return = np.dot(w_mv, mu)

print(f"  Optimization success: {result_mv.success}")
print(f"  Weights:")
for i, code in enumerate(stock_codes):
    print(f"    {code} ({stock_names[i]}): {w_mv[i]:.6f} ({w_mv[i]*100:.4f}%)")
print(f"  Expected Return: {mv_return:.6f} ({mv_return*100:.4f}%)")
print(f"  Standard Deviation: {mv_std:.6f} ({mv_std*100:.4f}%)")

# ---- Drawing the Efficient Frontier (no short selling) ----
print("\n  Drawing Efficient Frontier...")

# Find range of target returns for efficient frontier
# Minimum return is the MV portfolio return
# Maximum return is the highest expected return among stocks
min_ret = mv_return
max_ret = max(mu)

target_returns = np.linspace(min_ret, max_ret, 200)
ef_std = []
ef_ret = []
ef_weights_list = []

for target_ret in target_returns:
    constraints_ef = [
        {'type': 'eq', 'fun': lambda w: np.sum(w) - 1},
        {'type': 'eq', 'fun': lambda w, tr=target_ret: np.dot(w, mu) - tr}
    ]
    result_ef = minimize(portfolio_variance, w0, args=(Sigma,),
                         method='SLSQP', bounds=bounds, constraints=constraints_ef,
                         options={'ftol': 1e-15, 'maxiter': 1000})
    if result_ef.success:
        std_ef = np.sqrt(portfolio_variance(result_ef.x, Sigma))
        ef_std.append(std_ef)
        ef_ret.append(target_ret)
        ef_weights_list.append(result_ef.x)

# Plot the efficient frontier
plt.figure(figsize=(12, 8))
plt.plot(ef_std, ef_ret, 'b-', linewidth=2, label='Efficient Frontier (No Short Sale)')

# Plot individual stocks
stock_stds = [annualized_std[code] for code in stock_codes]
stock_rets = [expected_returns[code] for code in stock_codes]
plt.scatter(stock_stds, stock_rets, c='green', marker='o', s=100, zorder=5, label='Individual Stocks')
for i, code in enumerate(stock_codes):
    plt.annotate(f'{code}', (stock_stds[i], stock_rets[i]), textcoords="offset points",
                 xytext=(5, 5), fontsize=8)

# Plot EW portfolio
plt.scatter([ew_std], [ew_return], c='orange', marker='s', s=150, zorder=5, label=f'EW Portfolio')

# Plot MV portfolio
plt.scatter([mv_std], [mv_return], c='red', marker='^', s=150, zorder=5, label=f'MV Portfolio')

plt.xlabel('Standard Deviation (Annualized)')
plt.ylabel('Expected Return (Annualized)')
plt.title('Efficient Frontier (Short Sale Not Allowed)')
plt.legend()
plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('efficient_frontier.png', dpi=150, bbox_inches='tight')
print("  Efficient Frontier saved to 'efficient_frontier.png'")

# ---- Q2(i)(c): Optimal Portfolio (OP) - Tangency Portfolio ----
print("\n--- Q2(i)(c): Optimal Portfolio (OP) - Tangency Portfolio ---")

def neg_sharpe_ratio(w, mu, Sigma, rf):
    port_ret = np.dot(w, mu)
    port_std = np.sqrt(np.dot(w, np.dot(Sigma, w)))
    if port_std < 1e-10:
        return 1e10
    return -(port_ret - rf) / port_std

result_op = minimize(neg_sharpe_ratio, w0, args=(mu, Sigma, rf),
                     method='SLSQP', bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-15, 'maxiter': 1000})

w_op = result_op.x
op_return = np.dot(w_op, mu)
op_var = np.dot(w_op, np.dot(Sigma, w_op))
op_std = np.sqrt(op_var)
sharpe = (op_return - rf) / op_std

print(f"  Optimization success: {result_op.success}")
print(f"  Weights:")
for i, code in enumerate(stock_codes):
    print(f"    {code} ({stock_names[i]}): {w_op[i]:.6f} ({w_op[i]*100:.4f}%)")
print(f"  Expected Return: {op_return:.6f} ({op_return*100:.4f}%)")
print(f"  Standard Deviation: {op_std:.6f} ({op_std*100:.4f}%)")
print(f"  Sharpe Ratio: {sharpe:.6f}")

# Add OP and CML to the plot
plt.figure(figsize=(12, 8))
plt.plot(ef_std, ef_ret, 'b-', linewidth=2, label='Efficient Frontier (No Short Sale)')

# Capital Market Line
cml_std = np.linspace(0, max(ef_std) * 1.2, 100)
cml_ret = rf + sharpe * cml_std
plt.plot(cml_std, cml_ret, 'r--', linewidth=1.5, label='Capital Market Line')

# Plot risk-free asset
plt.scatter([0], [rf], c='black', marker='*', s=200, zorder=5, label=f'Risk-Free ({rf*100}%)')

# Plot individual stocks
plt.scatter(stock_stds, stock_rets, c='green', marker='o', s=100, zorder=5, label='Individual Stocks')
for i, code in enumerate(stock_codes):
    plt.annotate(f'{code}', (stock_stds[i], stock_rets[i]), textcoords="offset points",
                 xytext=(5, 5), fontsize=8)

# Plot EW, MV, OP
plt.scatter([ew_std], [ew_return], c='orange', marker='s', s=150, zorder=5, label=f'EW Portfolio')
plt.scatter([mv_std], [mv_return], c='red', marker='^', s=150, zorder=5, label=f'MV Portfolio')
plt.scatter([op_std], [op_return], c='purple', marker='D', s=150, zorder=5, label=f'OP (Tangency)')

plt.xlabel('Standard Deviation (Annualized)')
plt.ylabel('Expected Return (Annualized)')
plt.title('Efficient Frontier with CML (Short Sale Not Allowed)')
plt.legend(loc='lower right')
plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('efficient_frontier_with_CML.png', dpi=150, bbox_inches='tight')
print("  Efficient Frontier with CML saved to 'efficient_frontier_with_CML.png'")

# ============================================================================
# Q2(ii): Risk-Parity Strategy
# ============================================================================
print("\n" + "-" * 60)
print("Q2(ii): Risk-Parity Strategy")
print("-" * 60)

# ---- Q2(ii)(a): Equal Risk Contribution (ERC) Portfolio ----
print("\n--- Q2(ii)(a): Equal Risk Contribution (ERC) Portfolio ---")

def erc_objective(w, Sigma):
    """
    Objective: minimize sum of (w_i * (Sigma @ w)_i - w_j * (Sigma @ w)_j)^2
    for all pairs i,j. This makes all total risk contributions equal.
    
    Equivalently, minimize sum_i (w_i * (Sigma @ w)_i / sigma_p - 1/n)^2
    """
    n = len(w)
    sigma_w = Sigma @ w
    port_var = w @ sigma_w
    if port_var < 1e-15:
        return 1e10
    port_std = np.sqrt(port_var)
    
    # Total risk contribution of each asset
    trc = w * sigma_w / port_std
    
    # Target: each should contribute 1/n of total risk (port_std)
    target = port_std / n
    
    # Minimize sum of squared differences
    return np.sum((trc - target) ** 2)

# Alternative: use the log-barrier formulation for ERC
# Minimize: 0.5 * w' Sigma w - c * sum(ln(w_i))
# This is the Maillard, Roncalli, Teïlétché (2010) approach
def erc_log_objective(w, Sigma, c=1e-4):
    port_var = w @ Sigma @ w
    log_sum = np.sum(np.log(w + 1e-20))
    return 0.5 * port_var - c * log_sum

# Use direct approach: minimize differences in risk contributions
constraints_erc = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
bounds_erc = tuple((1e-6, 1) for _ in range(n_stocks))
w0_erc = np.array([1.0 / n_stocks] * n_stocks)

result_erc = minimize(erc_objective, w0_erc, args=(Sigma,),
                      method='SLSQP', bounds=bounds_erc, constraints=constraints_erc,
                      options={'ftol': 1e-20, 'maxiter': 5000})

w_erc = result_erc.x

# Normalize weights
w_erc = w_erc / np.sum(w_erc)

erc_return = np.dot(w_erc, mu)
erc_var = np.dot(w_erc, np.dot(Sigma, w_erc))
erc_std = np.sqrt(erc_var)

# Marginal risk contribution: (Sigma @ w) / sigma_p
sigma_w_erc = Sigma @ w_erc
mrc_erc = sigma_w_erc / erc_std

# Total risk contribution: w_i * MRC_i
trc_erc = w_erc * mrc_erc

print(f"  Optimization success: {result_erc.success}")
print(f"\n  {'Stock':<8} {'Weight':>10} {'MRC':>12} {'TRC':>12} {'TRC %':>10}")
print("  " + "-" * 55)
for i, code in enumerate(stock_codes):
    print(f"  {code:<8} {w_erc[i]:>10.6f} {mrc_erc[i]:>12.6f} {trc_erc[i]:>12.6f} {trc_erc[i]/erc_std*100:>9.4f}%")
print(f"\n  Sum of weights: {np.sum(w_erc):.6f}")
print(f"  Sum of TRC: {np.sum(trc_erc):.6f} (should equal sigma_p = {erc_std:.6f})")
print(f"  Expected Return: {erc_return:.6f} ({erc_return*100:.4f}%)")
print(f"  Standard Deviation: {erc_std:.6f} ({erc_std*100:.4f}%)")

# ---- Q2(ii)(b): Leveraged ERC Portfolio (ERC-L) ----
print("\n--- Q2(ii)(b): Leveraged ERC Portfolio (ERC-L) ---")

# Target standard deviation = EW portfolio standard deviation
target_std = ew_std
leverage = target_std / erc_std

# Expected return of ERC-L from Capital Allocation Line:
# E[r_ERC-L] = rf + leverage * (E[r_ERC] - rf)
erc_l_return = rf + leverage * (erc_return - rf)
erc_l_std = leverage * erc_std  # Should equal ew_std

print(f"  Target s.d. (= EW s.d.): {target_std:.6f} ({target_std*100:.4f}%)")
print(f"  ERC s.d.: {erc_std:.6f} ({erc_std*100:.4f}%)")
print(f"  Leverage factor: {leverage:.6f}")
print(f"  ERC-L Expected Return: {erc_l_return:.6f} ({erc_l_return*100:.4f}%)")
print(f"  ERC-L Standard Deviation: {erc_l_std:.6f} ({erc_l_std*100:.4f}%)")

# ============================================================================
# QUESTION 3: Presenting the Performance of the Portfolios (6 marks)
# ============================================================================
print("\n" + "=" * 80)
print("QUESTION 3: Portfolio Performance")
print("=" * 80)

# Stock prices on Feb 2, 2026 and March 2, 2026
prices_feb2 = {
    267: 12.20, 388: 424.20, 700: 598.00, 941: 78.70,
    981: 72.20, 1801: 77.80, 2319: 16.50, 2899: 39.56,
    9888: 141.40, 9988: 164.70
}
hsi_feb2 = 26776.0

prices_mar2 = {
    267: 12.25, 388: 411.00, 700: 511.50, 941: 78.50,
    981: 64.60, 1801: 82.60, 2319: 15.54, 2899: 46.04,
    9888: 118.50, 9988: 136.40
}
hsi_mar2 = 26059.85

print("\nStock prices:")
print(f"{'Stock':<8} {'Feb 2, 2026':>14} {'Mar 2, 2026':>14} {'Return':>10}")
print("-" * 48)
for code in stock_codes:
    ret = (prices_mar2[code] - prices_feb2[code]) / prices_feb2[code]
    print(f"  {code:<6} {prices_feb2[code]:>12.2f} {prices_mar2[code]:>12.2f} {ret*100:>9.4f}%")
hsi_ret = (hsi_mar2 - hsi_feb2) / hsi_feb2
print(f"  {'HSI':<6} {hsi_feb2:>12.2f} {hsi_mar2:>12.2f} {hsi_ret*100:>9.4f}%")

investment = 1_000_000  # $1,000,000

# Calculate portfolio values
portfolios = {
    'EW': w_ew,
    'MV': w_mv,
    'ERC': w_erc,
}

print(f"\nInvestment amount: ${investment:,.0f}")
print(f"\n{'Portfolio':<12} {'Initial':>14} {'Final':>14} {'Return':>10}")
print("-" * 52)

def calc_portfolio_value(weights, prices_start, prices_end, investment_amount):
    """Calculate portfolio value given weights and prices."""
    # For each stock, investment = weight * total_investment
    # Number of shares (fractional) = (weight * investment) / price_start
    # Value at end = shares * price_end
    total_value = 0
    for i, code in enumerate(stock_codes):
        amount_invested = weights[i] * investment_amount
        shares = amount_invested / prices_start[code]
        end_value = shares * prices_end[code]
        total_value += end_value
    return total_value

# EW portfolio
ew_final = calc_portfolio_value(w_ew, prices_feb2, prices_mar2, investment)
ew_perf = (ew_final - investment) / investment
print(f"  {'EW':<10} ${investment:>12,.2f} ${ew_final:>12,.2f} {ew_perf*100:>9.4f}%")

# MV portfolio
mv_final = calc_portfolio_value(w_mv, prices_feb2, prices_mar2, investment)
mv_perf = (mv_final - investment) / investment
print(f"  {'MV':<10} ${investment:>12,.2f} ${mv_final:>12,.2f} {mv_perf*100:>9.4f}%")

# 0.7*OP + 0.3*rf portfolio
# The risky portion (70%) is invested in OP, 30% in risk-free asset
risky_investment = 0.7 * investment
rf_investment = 0.3 * investment

op_risky_final = calc_portfolio_value(w_op, prices_feb2, prices_mar2, risky_investment)
# Risk-free portion: grows at rf rate for ~28 days (Feb 2 to Mar 2)
# Since "no compounding required", use simple return
# 28 calendar days / 365 * rf
days_period = 28  # Feb 2 to Mar 2
rf_period_return = rf * days_period / 365
rf_final = rf_investment * (1 + rf_period_return)
op_combined_final = op_risky_final + rf_final
op_combined_perf = (op_combined_final - investment) / investment
print(f"  {'0.7OP+0.3r':<10} ${investment:>12,.2f} ${op_combined_final:>12,.2f} {op_combined_perf*100:>9.4f}%")

# ERC portfolio
erc_final = calc_portfolio_value(w_erc, prices_feb2, prices_mar2, investment)
erc_perf = (erc_final - investment) / investment
print(f"  {'ERC':<10} ${investment:>12,.2f} ${erc_final:>12,.2f} {erc_perf*100:>9.4f}%")

# ERC-L portfolio
# Leveraged portfolio: borrow (leverage - 1) * investment at rf, invest leverage * investment in ERC
erc_l_risky_investment = leverage * investment
erc_l_borrow = (leverage - 1) * investment
erc_l_risky_final = calc_portfolio_value(w_erc, prices_feb2, prices_mar2, erc_l_risky_investment)
erc_l_borrow_cost = erc_l_borrow * (1 + rf_period_return)
erc_l_final = erc_l_risky_final - erc_l_borrow_cost + investment  # net value
# Actually: final = leverage * ERC_return_on_investment - (leverage-1) * rf_cost + investment
# Simpler: ERC-L return = leverage * ERC_return - (leverage - 1) * rf_period_return
erc_l_actual_return = leverage * erc_perf - (leverage - 1) * rf_period_return
erc_l_final_v2 = investment * (1 + erc_l_actual_return)
print(f"  {'ERC-L':<10} ${investment:>12,.2f} ${erc_l_final_v2:>12,.2f} {erc_l_actual_return*100:>9.4f}%")

# HSI
print(f"  {'HSI':<10} ${investment:>12,.2f} ${investment*(1+hsi_ret):>12,.2f} {hsi_ret*100:>9.4f}%")

# ============================================================================
# Summary Table
# ============================================================================
print("\n" + "=" * 80)
print("SUMMARY OF RESULTS")
print("=" * 80)

print("\n--- Portfolio Weights ---")
print(f"{'Stock':<8} {'EW':>8} {'MV':>8} {'OP':>8} {'ERC':>8}")
print("-" * 40)
for i, code in enumerate(stock_codes):
    print(f"  {code:<6} {w_ew[i]:>7.4f} {w_mv[i]:>7.4f} {w_op[i]:>7.4f} {w_erc[i]:>7.4f}")
print(f"  {'Sum':<6} {sum(w_ew):>7.4f} {sum(w_mv):>7.4f} {sum(w_op):>7.4f} {sum(w_erc):>7.4f}")

print(f"\n--- Portfolio Expected Return & Standard Deviation ---")
print(f"{'Portfolio':<12} {'E[r]':>10} {'s.d.':>10} {'Sharpe':>10}")
print("-" * 44)
print(f"  {'EW':<10} {ew_return*100:>9.4f}% {ew_std*100:>9.4f}%")
print(f"  {'MV':<10} {mv_return*100:>9.4f}% {mv_std*100:>9.4f}%")
print(f"  {'OP':<10} {op_return*100:>9.4f}% {op_std*100:>9.4f}% {sharpe:>9.4f}")
print(f"  {'ERC':<10} {erc_return*100:>9.4f}% {erc_std*100:>9.4f}%")
print(f"  {'ERC-L':<10} {erc_l_return*100:>9.4f}% {erc_l_std*100:>9.4f}%")

print(f"\n--- Q3: Portfolio Performance (Feb 2 to Mar 2, 2026) ---")
print(f"{'Portfolio':<12} {'Return (%)':>12}")
print("-" * 26)
print(f"  {'EW':<10} {ew_perf*100:>10.4f}%")
print(f"  {'MV':<10} {mv_perf*100:>10.4f}%")
print(f"  {'0.7OP+0.3r':<10} {op_combined_perf*100:>10.4f}%")
print(f"  {'ERC':<10} {erc_perf*100:>10.4f}%")
print(f"  {'ERC-L':<10} {erc_l_actual_return*100:>10.4f}%")
print(f"  {'HSI':<10} {hsi_ret*100:>10.4f}%")

print("\n" + "=" * 80)
print("END OF ASSIGNMENT 2 SOLUTION")
print("=" * 80)
