"""
Save key results from assignment2_solution to a file
"""
import numpy as np
import pandas as pd
from scipy.optimize import minimize
import openpyxl
import warnings
warnings.filterwarnings('ignore')

wb = openpyxl.load_workbook('CMSC 5718 Assignment 2 parameters.xlsx', data_only=True)

# Load stock prices
ws_prices = wb['stock prices']
stock_codes = [267, 388, 700, 941, 981, 1801, 2319, 2899, 9888, 9988]
stock_names = ['Citic Limited', 'HK Exchanges', 'Tencent', 'China Mobile',
               'SMIC', 'Innovent Bio', 'Mengniu Dairy', 'Zijin Mining',
               'Baidu', 'Alibaba']

dates = []
price_data = {code: [] for code in stock_codes}
hsi_prices = []

for i, row in enumerate(ws_prices.iter_rows(values_only=True), 1):
    if i < 4:
        continue
    row_list = list(row)
    date_val = row_list[1]
    if date_val is None:
        continue
    if hasattr(date_val, 'strftime'):
        dates.append(date_val)
    else:
        from datetime import datetime
        try:
            dates.append(datetime.strptime(str(date_val), '%Y-%m-%d'))
        except:
            continue
    for j, code in enumerate(stock_codes):
        price_data[code].append(float(row_list[j + 2]))
    hsi_prices.append(float(row_list[12]))

price_df = pd.DataFrame(price_data, index=dates)
price_df['HSI'] = hsi_prices

# Expected returns
ws_er = wb['Consensus Expected return']
expected_returns = {}
for row in ws_er.iter_rows(values_only=True):
    row_list = list(row)
    if row_list[2] is not None and isinstance(row_list[2], (int, float)):
        code = int(row_list[2])
        if code in stock_codes:
            expected_returns[code] = float(row_list[4])

rf = 0.029
returns_df = price_df.pct_change().dropna()
n = len(returns_df)
trading_days = 252
daily_std = returns_df.std(ddof=1)
annualized_std = daily_std * np.sqrt(trading_days)

stock_returns = returns_df[stock_codes]
daily_cov = stock_returns.cov()
annualized_cov = daily_cov * trading_days

mu = np.array([expected_returns[code] for code in stock_codes])
Sigma = annualized_cov.values
n_stocks = len(stock_codes)

# Store results
results = {}

# Q1
results['n_returns'] = n
results['annualized_std'] = {code: annualized_std[code] for code in stock_codes}
results['annualized_std']['HSI'] = annualized_std['HSI']
results['annualized_cov'] = annualized_cov

# Q2(i)(a) EW
w_ew = np.array([0.1] * n_stocks)
ew_return = np.dot(w_ew, mu)
ew_var = np.dot(w_ew, np.dot(Sigma, w_ew))
ew_std = np.sqrt(ew_var)

# Q2(i)(b) MV
def portfolio_variance(w, Sigma):
    return np.dot(w, np.dot(Sigma, w))

constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
bounds = tuple((0, 1) for _ in range(n_stocks))
w0 = np.array([1.0 / n_stocks] * n_stocks)
result_mv = minimize(portfolio_variance, w0, args=(Sigma,),
                     method='SLSQP', bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-15, 'maxiter': 1000})
w_mv = result_mv.x
mv_std = np.sqrt(portfolio_variance(w_mv, Sigma))
mv_return = np.dot(w_mv, mu)

# Q2(i)(c) OP
def neg_sharpe_ratio(w, mu, Sigma, rf):
    port_ret = np.dot(w, mu)
    port_std = np.sqrt(np.dot(w, np.dot(Sigma, w)))
    if port_std < 1e-10:
        return 1e10
    return -(port_ret - rf) / port_std

result_op = minimize(neg_sharpe_ratio, w0, args=(mu, Sigma, rf),
                     method='SLSQP', bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-15, 'maxiter': 1000})
w_op = result_op.x
op_return = np.dot(w_op, mu)
op_std = np.sqrt(np.dot(w_op, np.dot(Sigma, w_op)))
sharpe = (op_return - rf) / op_std

# Q2(ii)(a) ERC
def erc_objective(w, Sigma):
    n_s = len(w)
    sigma_w = Sigma @ w
    port_var = w @ sigma_w
    if port_var < 1e-15:
        return 1e10
    port_std = np.sqrt(port_var)
    trc = w * sigma_w / port_std
    target = port_std / n_s
    return np.sum((trc - target) ** 2)

constraints_erc = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
bounds_erc = tuple((1e-6, 1) for _ in range(n_stocks))
result_erc = minimize(erc_objective, w0, args=(Sigma,),
                      method='SLSQP', bounds=bounds_erc, constraints=constraints_erc,
                      options={'ftol': 1e-20, 'maxiter': 5000})
w_erc = result_erc.x
w_erc = w_erc / np.sum(w_erc)
erc_return = np.dot(w_erc, mu)
erc_var = np.dot(w_erc, np.dot(Sigma, w_erc))
erc_std = np.sqrt(erc_var)
sigma_w_erc = Sigma @ w_erc
mrc_erc = sigma_w_erc / erc_std
trc_erc = w_erc * mrc_erc

# Q2(ii)(b) ERC-L
leverage = ew_std / erc_std
erc_l_return = rf + leverage * (erc_return - rf)
erc_l_std = leverage * erc_std

# Write results
with open('results_summary.md', 'w', encoding='utf-8') as f:
    f.write("# CMSC 5718 Assignment 2 - Results Summary\n")
    f.write("## Student Number: 1155239333\n\n")
    
    f.write("## Q1(i): Annualized Standard Deviation\n\n")
    f.write("| Stock | Ann. Std Dev |\n")
    f.write("|-------|-------------|\n")
    for code in stock_codes:
        f.write(f"| {code} | {annualized_std[code]:.6f} ({annualized_std[code]*100:.4f}%) |\n")
    f.write(f"| HSI | {annualized_std['HSI']:.6f} ({annualized_std['HSI']*100:.4f}%) |\n")
    
    f.write("\n## Q1(ii): Annualized Covariance Matrix\n\n")
    f.write("| | " + " | ".join(str(c) for c in stock_codes) + " |\n")
    f.write("|---" * (n_stocks + 1) + "|\n")
    for i, ci in enumerate(stock_codes):
        row = f"| {ci} |"
        for j, cj in enumerate(stock_codes):
            row += f" {annualized_cov.iloc[i, j]:.6f} |"
        f.write(row + "\n")
    
    f.write("\n## Q2(i)(a): Equal Weight Portfolio (EW)\n\n")
    f.write(f"- Expected Return: {ew_return:.6f} ({ew_return*100:.4f}%)\n")
    f.write(f"- Standard Deviation: {ew_std:.6f} ({ew_std*100:.4f}%)\n")
    
    f.write("\n## Q2(i)(b): Global Minimum Variance Portfolio (MV)\n\n")
    f.write("| Stock | Weight |\n")
    f.write("|-------|--------|\n")
    for i, code in enumerate(stock_codes):
        f.write(f"| {code} ({stock_names[i]}) | {w_mv[i]:.6f} ({w_mv[i]*100:.4f}%) |\n")
    f.write(f"\n- Expected Return: {mv_return:.6f} ({mv_return*100:.4f}%)\n")
    f.write(f"- Standard Deviation: {mv_std:.6f} ({mv_std*100:.4f}%)\n")
    
    f.write("\n## Q2(i)(c): Optimal Portfolio (OP / Tangency)\n\n")
    f.write("| Stock | Weight |\n")
    f.write("|-------|--------|\n")
    for i, code in enumerate(stock_codes):
        f.write(f"| {code} ({stock_names[i]}) | {w_op[i]:.6f} ({w_op[i]*100:.4f}%) |\n")
    f.write(f"\n- Expected Return: {op_return:.6f} ({op_return*100:.4f}%)\n")
    f.write(f"- Standard Deviation: {op_std:.6f} ({op_std*100:.4f}%)\n")
    f.write(f"- Sharpe Ratio: {sharpe:.6f}\n")
    
    f.write("\n## Q2(ii)(a): Equal Risk Contribution Portfolio (ERC)\n\n")
    f.write("| Stock | Weight | MRC | TRC | TRC% |\n")
    f.write("|-------|--------|-----|-----|------|\n")
    for i, code in enumerate(stock_codes):
        f.write(f"| {code} | {w_erc[i]:.6f} | {mrc_erc[i]:.6f} | {trc_erc[i]:.6f} | {trc_erc[i]/erc_std*100:.4f}% |\n")
    f.write(f"\n- Expected Return: {erc_return:.6f} ({erc_return*100:.4f}%)\n")
    f.write(f"- Standard Deviation: {erc_std:.6f} ({erc_std*100:.4f}%)\n")
    
    f.write("\n## Q2(ii)(b): Leveraged ERC Portfolio (ERC-L)\n\n")
    f.write(f"- Leverage Factor: {leverage:.6f}\n")
    f.write(f"- ERC-L Expected Return: {erc_l_return:.6f} ({erc_l_return*100:.4f}%)\n")
    f.write(f"- ERC-L Standard Deviation: {erc_l_std:.6f} ({erc_l_std*100:.4f}%)\n")
    
    # Q3
    prices_feb2 = {267: 12.20, 388: 424.20, 700: 598.00, 941: 78.70,
                   981: 72.20, 1801: 77.80, 2319: 16.50, 2899: 39.56,
                   9888: 141.40, 9988: 164.70}
    hsi_feb2 = 26776.0
    prices_mar2 = {267: 12.25, 388: 411.00, 700: 511.50, 941: 78.50,
                   981: 64.60, 1801: 82.60, 2319: 15.54, 2899: 46.04,
                   9888: 118.50, 9988: 136.40}
    hsi_mar2 = 26059.85
    
    investment = 1_000_000
    days_period = 28
    rf_period_return = rf * days_period / 365
    
    def calc_pv(weights, ps, pe, inv):
        total = 0
        for i_s, code in enumerate(stock_codes):
            shares = (weights[i_s] * inv) / ps[code]
            total += shares * pe[code]
        return total
    
    ew_f = calc_pv(w_ew, prices_feb2, prices_mar2, investment)
    mv_f = calc_pv(w_mv, prices_feb2, prices_mar2, investment)
    op_risky = calc_pv(w_op, prices_feb2, prices_mar2, 0.7 * investment)
    rf_f = 0.3 * investment * (1 + rf_period_return)
    op_cf = op_risky + rf_f
    erc_f = calc_pv(w_erc, prices_feb2, prices_mar2, investment)
    
    ew_p = (ew_f - investment) / investment
    mv_p = (mv_f - investment) / investment
    op_cp = (op_cf - investment) / investment
    erc_p = (erc_f - investment) / investment
    erc_l_p = leverage * erc_p - (leverage - 1) * rf_period_return
    hsi_r = (hsi_mar2 - hsi_feb2) / hsi_feb2
    
    f.write("\n## Q3: Portfolio Performance (Feb 2 - Mar 2, 2026)\n\n")
    f.write("### Stock Prices\n\n")
    f.write("| Stock | Feb 2 | Mar 2 | Return |\n")
    f.write("|-------|-------|-------|--------|\n")
    for code in stock_codes:
        r = (prices_mar2[code] - prices_feb2[code]) / prices_feb2[code]
        f.write(f"| {code} | {prices_feb2[code]:.2f} | {prices_mar2[code]:.2f} | {r*100:.4f}% |\n")
    f.write(f"| HSI | {hsi_feb2:.2f} | {hsi_mar2:.2f} | {hsi_r*100:.4f}% |\n")
    
    f.write("\n### Portfolio Performance\n\n")
    f.write("| Portfolio | Initial Value | Final Value | Return |\n")
    f.write("|-----------|---------------|-------------|--------|\n")
    f.write(f"| EW | $1,000,000 | ${ew_f:,.2f} | {ew_p*100:.4f}% |\n")
    f.write(f"| MV | $1,000,000 | ${mv_f:,.2f} | {mv_p*100:.4f}% |\n")
    f.write(f"| 0.7OP+0.3r | $1,000,000 | ${op_cf:,.2f} | {op_cp*100:.4f}% |\n")
    f.write(f"| ERC | $1,000,000 | ${erc_f:,.2f} | {erc_p*100:.4f}% |\n")
    erc_l_fv = investment * (1 + erc_l_p)
    f.write(f"| ERC-L | $1,000,000 | ${erc_l_fv:,.2f} | {erc_l_p*100:.4f}% |\n")
    f.write(f"| HSI | $1,000,000 | ${investment*(1+hsi_r):,.2f} | {hsi_r*100:.4f}% |\n")

print("Results saved to results_summary.md")

# Also generate the efficient frontier plot
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Efficient frontier
min_ret = mv_return
max_ret = max(mu)
target_returns = np.linspace(min_ret, max_ret, 200)
ef_std = []
ef_ret = []

for target_ret in target_returns:
    constraints_ef = [
        {'type': 'eq', 'fun': lambda w: np.sum(w) - 1},
        {'type': 'eq', 'fun': lambda w, tr=target_ret: np.dot(w, mu) - tr}
    ]
    result_ef = minimize(portfolio_variance, w0, args=(Sigma,),
                         method='SLSQP', bounds=bounds, constraints=constraints_ef,
                         options={'ftol': 1e-15, 'maxiter': 1000})
    if result_ef.success:
        std_ef = np.sqrt(portfolio_variance(result_ef.x, Sigma))
        ef_std.append(std_ef)
        ef_ret.append(target_ret)

fig, ax = plt.subplots(figsize=(12, 8))
ax.plot(ef_std, ef_ret, 'b-', linewidth=2, label='Efficient Frontier')

# CML
cml_std_arr = np.linspace(0, max(ef_std) * 1.2, 100)
cml_ret_arr = rf + sharpe * cml_std_arr
ax.plot(cml_std_arr, cml_ret_arr, 'r--', linewidth=1.5, label='Capital Market Line')

# Risk-free
ax.scatter([0], [rf], c='black', marker='*', s=200, zorder=5, label=f'Risk-Free ({rf*100}%)')

# Individual stocks
s_stds = [annualized_std[code] for code in stock_codes]
s_rets = [expected_returns[code] for code in stock_codes]
ax.scatter(s_stds, s_rets, c='green', marker='o', s=100, zorder=5, label='Individual Stocks')
for i_s, code in enumerate(stock_codes):
    ax.annotate(f'{code}', (s_stds[i_s], s_rets[i_s]), textcoords="offset points",
                xytext=(5, 5), fontsize=8)

ax.scatter([ew_std], [ew_return], c='orange', marker='s', s=150, zorder=5, label='EW')
ax.scatter([mv_std], [mv_return], c='red', marker='^', s=150, zorder=5, label='MV')
ax.scatter([op_std], [op_return], c='purple', marker='D', s=150, zorder=5, label='OP')
ax.scatter([erc_std], [erc_return], c='cyan', marker='P', s=150, zorder=5, label='ERC')

ax.set_xlabel('Standard Deviation (Annualized)')
ax.set_ylabel('Expected Return (Annualized)')
ax.set_title('Efficient Frontier with CML (Short Sale Not Allowed)')
ax.legend(loc='lower right')
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('efficient_frontier.png', dpi=150, bbox_inches='tight')
print("Efficient frontier plot saved to efficient_frontier.png")
